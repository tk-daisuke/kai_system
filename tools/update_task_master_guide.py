# -*- coding: utf-8 -*-
"""
Task_Master.xlsx に「設定方法」シートを追加・更新するスクリプト
"""

import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def find_task_master_files():
    """settingsフォルダ以下のTask_Master.xlsxを探す"""
    base_path = Path(__file__).parent.parent
    settings_dir = base_path / "settings"
    
    files = list(settings_dir.rglob("Task_Master.xlsx"))
    return files

def create_guide_sheet(file_path):
    """指定されたファイルにガイドシートを追加・更新"""
    print(f"Processing: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return

    sheet_name = "設定方法"
    
    # 既存シートがあれば削除して作り直す（最新化のため）
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # ガイドデータ
    headers = ["項目名 (日本語 / 英語)", "説明", "設定例", "備考"]
    data = [
        ["グループ / Group", "GUIのボタンとしてグループ化するための名称。同じグループは連続実行されます。", "朝処理", "必須"],
        ["開始時間 / StartTime", "タスクを実行可能な開始時刻 (HH:MM)。この時間より前には実行されません。", "09:00", "必須"],
        ["終了時間 / EndTime", "タスク実行可能な終了時刻。空欄の場合は開始+8時間とみなされます。", "18:00", "任意"],
        ["ファイルパス / FilePath", "操作対象となるExcelファイルの絶対パス。", "C:\\Work\\Data.xlsx", "必須"],
        ["転記シート / TargetSheet", "ダウンロードしたCSVデータを貼り付けるシート名。", "Sheet1", "DLスキップ時は無視"],
        ["検索キー / SearchKey", "ダウンロードフォルダ内で検索するCSVファイル名の一部（キーワード）。最新のものを対象とします。", "user_list", "DLスキップ時は無視"],
        ["ダウンロードURL / DownloadURL", "データ取得元のURL。ブラウザで開くか、直接DLします。", "https://example.com/dl", "DLスキップ時は無視"],
        ["完了後動作 / ActionAfter", "処理完了後の動作。\n'Save': 保存して閉じる/次へ\n'Pause': メッセージを表示して一時停止（手動作業待ち）", "Save", "デフォルト: Save"],
        ["有効 / Active", "タスクを有効にするか。TRUE または 1 で有効。", "TRUE", "デフォルト: FALSE"],
        ["DLスキップ / SkipDownload", "ダウンロードと転記処理をスキップし、ファイルを開くだけにします。", "FALSE", "デフォルト: FALSE"],
        ["終了後閉じる / CloseAfter", "タスク完了後にファイルを閉じるかどうか。\nTRUE: 閉じる\nFALSE: 開いたまま（連続処理でファイルを再利用するため）", "FALSE", "デフォルト: FALSE (開いたまま)"],
        ["メッセージ / PopupMessage", "ActionAfterが 'Pause' の時に表示するカスタムメッセージ。\n改行コードも使用可能。", "データの更新を確認して\nOKを押してください", "任意"],
        ["マクロ名 / MacroName", "処理後に実行したいVBAマクロ名。", "RunAnalysis", "任意（未実装機能の可能性あり）"]
    ]

    # ヘッダー書き込み
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # データ書き込み
    row_font = Font(name="Yu Gothic UI", size=11)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.alignment = wrap_align
            cell.border = border
            
            # 項目名の列は少し太字で見やすく
            if col_idx == 1:
                cell.font = Font(name="Yu Gothic UI", size=11, bold=True)

    # 列幅調整
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    
    # 最初のシート（TaskList）をアクティブに戻す
    if "TaskList" in wb.sheetnames:
        idx = wb.sheetnames.index("TaskList")
        wb.active = idx

    wb.save(file_path)
    print(f"Updated: {file_path}")

if __name__ == "__main__":
    target_files = find_task_master_files()
    if not target_files:
        print("No Task_Master.xlsx found.")
    else:
        for f in target_files:
            create_guide_sheet(f)
