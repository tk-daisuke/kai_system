#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Task_Master.xlsx の構造を改善するスクリプト

- 列の整理（不要列削除、新列追加）
- 入力規則の設定
- セルの書式設定
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from pathlib import Path


# 新しい列構成の定義
NEW_COLUMNS = [
    {
        "name": "有効",
        "width": 8,
        "validation": "TRUE,FALSE",
        "comment": "TRUE: タスクを実行\nFALSE: スキップ",
        "default": "FALSE"
    },
    {
        "name": "グループ",
        "width": 15,
        "comment": "タスクをグループ化する名前\n例: 朝処理, 13:00業務"
    },
    {
        "name": "タスク名",
        "width": 20,
        "comment": "タスクの説明（任意）\n例: 売上データ取得"
    },
    {
        "name": "開始時刻",
        "width": 10,
        "comment": "実行開始時刻（HH:MM）\nこの時刻より前は待機します"
    },
    {
        "name": "終了時刻",
        "width": 10,
        "comment": "実行終了時刻（HH:MM）\n空欄の場合は開始+8時間"
    },
    {
        "name": "ファイルパス",
        "width": 50,
        "comment": "操作対象のExcelファイルの絶対パス\n例: C:\\Work\\Data.xlsx"
    },
    {
        "name": "シート",
        "width": 15,
        "comment": "CSVデータを貼り付けるシート名"
    },
    {
        "name": "URL",
        "width": 50,
        "comment": "CSVダウンロードURL"
    },
    {
        "name": "完了後動作",
        "width": 12,
        "validation": "None,Pause,Save",
        "comment": "None: 何もしない\nPause: 確認後に保存\nSave: 自動保存",
        "default": "Save"
    },
    {
        "name": "マクロ名",
        "width": 20,
        "comment": "実行するVBAマクロ名（任意）\n例: Module1.ImportData"
    },
    {
        "name": "DLスキップ",
        "width": 12,
        "validation": "TRUE,FALSE",
        "comment": "TRUE: ダウンロードせずファイルを開くのみ",
        "default": "FALSE"
    },
    {
        "name": "閉じる",
        "width": 10,
        "validation": "TRUE,FALSE",
        "comment": "TRUE: タスク完了後にExcelを閉じる\nFALSE: 開いたまま（連続処理用）",
        "default": "FALSE"
    },
    {
        "name": "曜日",
        "width": 15,
        "comment": "実行する曜日（カンマ区切り）\n1=月, 2=火, 3=水, 4=木, 5=金, 6=土, 7=日\n例: 1,2,3,4,5（平日のみ）\n空欄=毎日"
    },
    {
        "name": "祝日スキップ",
        "width": 12,
        "validation": "TRUE,FALSE",
        "comment": "TRUE: 日本の祝日はスキップ",
        "default": "FALSE"
    },
    {
        "name": "日付条件",
        "width": 15,
        "comment": "実行する日（カンマ区切り）\n例: 1,15（毎月1日と15日）\nL=月末\n空欄=毎日"
    },
    {
        "name": "メモ",
        "width": 30,
        "comment": "自由記述のメモ欄"
    },
]


def create_new_task_master(file_path: Path, preserve_data: bool = True):
    """
    新しい構造のTask_Master.xlsxを作成
    
    Args:
        file_path: 対象ファイルのパス
        preserve_data: 既存データを保持するか
    """
    print(f"処理中: {file_path}")
    
    # 既存データの読み込み
    old_data = []
    old_headers = []
    if file_path.exists() and preserve_data:
        try:
            old_wb = openpyxl.load_workbook(file_path)
            if "TaskList" in old_wb.sheetnames:
                old_ws = old_wb["TaskList"]
                old_headers = [cell.value for cell in old_ws[1]]
                for row in old_ws.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        old_data.append(dict(zip(old_headers, row)))
                print(f"  既存データ: {len(old_data)} 行を読み込み")
            old_wb.close()
        except Exception as e:
            print(f"  既存データ読み込みエラー: {e}")
    
    # 新しいワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TaskList"
    
    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF", name="Yu Gothic UI", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ヘッダー行を書き込み
    for col_idx, col_def in enumerate(NEW_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        
        # 列幅設定
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]
        
        # コメント追加
        if "comment" in col_def:
            cell.comment = Comment(col_def["comment"], "Co-worker Bot")
    
    # ヘッダー行を固定
    ws.freeze_panes = "A2"
    
    # 入力規則の設定
    for col_idx, col_def in enumerate(NEW_COLUMNS, 1):
        if "validation" in col_def:
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="list",
                formula1=f'"{col_def["validation"]}"',
                allow_blank=True,
                showDropDown=False
            )
            dv.error = f"無効な値です。{col_def['validation']} から選択してください。"
            dv.errorTitle = "入力エラー"
            dv.prompt = col_def.get("comment", "")
            dv.promptTitle = col_def["name"]
            dv.add(f"{col_letter}2:{col_letter}200")
            ws.add_data_validation(dv)
    
    # 列名のマッピング（旧列名 → 新列名）
    column_mapping = {
        "有効": "有効",
        "Active": "有効",
        "グループ": "グループ",
        "Group": "グループ",
        "Memo": "メモ",
        "メモ": "メモ",
        "開始時刻": "開始時刻",
        "StartTime": "開始時刻",
        "終了時刻": "終了時刻",
        "EndTime": "終了時刻",
        "ファイルパス": "ファイルパス",
        "FilePath": "ファイルパス",
        "転記シート": "シート",
        "CSV転記シート": "シート",
        "TargetSheet": "シート",
        "シート": "シート",
        "URL": "URL",
        "ダウンロードURL": "URL",
        "DownloadURL": "URL",
        "完了後動作": "完了後動作",
        "ActionAfter": "完了後動作",
        "マクロ名": "マクロ名",
        "MacroName": "マクロ名",
        "DLスキップ": "DLスキップ",
        "SkipDownload": "DLスキップ",
        "終了後閉じる": "閉じる",
        "CloseAfter": "閉じる",
        "閉じる": "閉じる",
        "曜日": "曜日",
        "Weekdays": "曜日",
        "祝日スキップ": "祝日スキップ",
        "SkipHoliday": "祝日スキップ",
        "日付条件": "日付条件",
        "DateCondition": "日付条件",
        # 検索キーは無視（廃止）
    }
    
    new_col_names = [c["name"] for c in NEW_COLUMNS]
    
    # 既存データを新しい構造にマッピング
    data_start_row = 2
    for row_idx, old_row in enumerate(old_data, data_start_row):
        for col_idx, col_def in enumerate(NEW_COLUMNS, 1):
            new_col_name = col_def["name"]
            
            # 旧データから値を探す
            value = None
            for old_col_name, mapped_name in column_mapping.items():
                if mapped_name == new_col_name and old_col_name in old_row:
                    value = old_row[old_col_name]
                    break
            
            # デフォルト値の設定
            if value is None and "default" in col_def:
                value = col_def["default"]
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # サンプル行の追加（既存データがない場合）
    if not old_data:
        sample_data = [
            {
                "有効": False,
                "グループ": "08:00業務",
                "タスク名": "朝の売上データ取得",
                "開始時刻": "08:00",
                "終了時刻": "09:00",
                "ファイルパス": "C:\\Work\\Sales.xlsx",
                "シート": "データ",
                "URL": "https://example.com/download/sales.csv",
                "完了後動作": "Save",
                "マクロ名": "",
                "DLスキップ": False,
                "閉じる": False,
                "曜日": "1,2,3,4,5",
                "祝日スキップ": True,
                "日付条件": "",
                "メモ": "サンプル: 平日のみ実行",
            }
        ]
        
        for row_idx, sample_row in enumerate(sample_data, 2):
            for col_idx, col_def in enumerate(NEW_COLUMNS, 1):
                value = sample_row.get(col_def["name"], col_def.get("default", ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
    
    # 「設定方法」シートを追加
    create_guide_sheet(wb)
    
    # 保存
    wb.save(file_path)
    print(f"  保存しました: {file_path.name}")


def create_guide_sheet(wb):
    """設定方法シートを作成"""
    sheet_name = "設定方法"
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # ヘッダー
    headers = ["項目名", "説明", "設定例", "備考"]
    
    # データ
    guide_data = [
        ["有効", "タスクを有効にするか", "TRUE", "必須"],
        ["グループ", "GUIのボタン名。同じグループは連続実行", "08:00業務", "必須"],
        ["タスク名", "タスクの説明（GUIには表示されない）", "売上データ取得", "任意"],
        ["開始時刻", "実行開始時刻。この時刻より前は待機", "09:00", "必須"],
        ["終了時刻", "実行終了時刻。過ぎたらスキップ", "18:00", "空欄=開始+8時間"],
        ["ファイルパス", "操作対象Excelファイルの絶対パス", "C:\\Work\\Data.xlsx", "必須"],
        ["シート", "CSVデータ貼り付け先シート名", "Sheet1", "DLスキップ時は不要"],
        ["URL", "CSVダウンロードURL", "https://example.com/dl", "DLスキップ時は不要"],
        ["完了後動作", "処理後の動作", "Save", "None/Pause/Save"],
        ["マクロ名", "実行するVBAマクロ", "Module1.Run", "任意"],
        ["DLスキップ", "ダウンロードせずファイルを開くのみ", "FALSE", "デフォルト: FALSE"],
        ["閉じる", "完了後にExcelを閉じる", "FALSE", "デフォルト: FALSE"],
        ["曜日", "実行曜日 (1=月〜7=日)", "1,2,3,4,5", "空欄=毎日"],
        ["祝日スキップ", "日本の祝日をスキップ", "TRUE", "デフォルト: FALSE"],
        ["日付条件", "実行日 (1,15=毎月1日と15日, L=月末)", "1,15,L", "空欄=毎日"],
        ["メモ", "自由記述欄", "〇〇システム用", "任意"],
    ]
    
    # スタイル
    header_font = Font(bold=True, color="FFFFFF", name="Yu Gothic UI", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ヘッダー書き込み
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # データ書き込み
    for row_idx, row_data in enumerate(guide_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Yu Gothic UI", size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if col_idx == 1:
                cell.font = Font(name="Yu Gothic UI", size=11, bold=True)
    
    # 列幅
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    
    # TaskListをアクティブに
    wb.active = wb.sheetnames.index("TaskList")


if __name__ == "__main__":
    base_path = Path(__file__).parent.parent  # toolsの親ディレクトリ（プロジェクトルート）
    
    # テスト設定ファイルを先に処理（確認用）
    test_file = base_path / "settings" / "test" / "Task_Master.xlsx"
    if test_file.exists():
        create_new_task_master(test_file, preserve_data=True)
    else:
        print(f"テストファイルが見つかりません: {test_file}")
    
    # 本番設定ファイルも処理
    prod_file = base_path / "settings" / "production" / "Task_Master.xlsx"
    if prod_file.exists():
        create_new_task_master(prod_file, preserve_data=True)
    else:
        print(f"本番ファイルが見つかりません: {prod_file}")
    
    print("\n完了!")
