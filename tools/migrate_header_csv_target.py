import pandas as pd
from pathlib import Path
import openpyxl

def migrate_headers_v2():
    # settingsフォルダを探す
    base_path = Path(__file__).parent.parent
    target_file = base_path / "settings" / "production" / "Task_Master.xlsx"
    
    if not target_file.exists():
        # フォールバック
        target_file = base_path / "settings" / "Task_Master.xlsx"
        
    if not target_file.exists():
        print(f"File not found: {target_file}")
        return

    print(f"Target: {target_file}")

    # マッピング定義（転記シート -> CSV転記シート）
    mapping = {
        "転記シート": "CSV転記シート",
        "TargetSheet": "CSV転記シート" # 念のため英語名からも
    }

    try:
        wb = openpyxl.load_workbook(target_file)
        if "TaskList" in wb.sheetnames:
            ws = wb["TaskList"]
            
            # ヘッダー行(1行目)を走査して置換
            for cell in ws[1]:
                if cell.value in mapping:
                    print(f"Renaming {cell.value} -> {mapping[cell.value]}")
                    cell.value = mapping[cell.value]
            
            wb.save(target_file)
            print("Migration (v2) completed successfully.")
        else:
            print("Sheet 'TaskList' not found.")
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    migrate_headers_v2()
