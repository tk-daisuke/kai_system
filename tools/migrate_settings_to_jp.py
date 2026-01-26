import pandas as pd
from pathlib import Path
import openpyxl

def migrate_headers():
    # settingsフォルダを探す
    base_path = Path(__file__).parent.parent
    target_file = base_path / "settings" / "production" / "Task_Master.xlsx"
    
    if not target_file.exists():
        # フォールバック
        target_file = base_path / "settings" / "Task_Master.xlsx"
        
    if not target_file.exists():
        print(f"File not found: {target_file}")
        return

    print(f"Target: {target_file}")

    # マッピング定義
    mapping = {
        "Active": "有効",
        "Group": "グループ",
        "StartTime": "開始時刻",
        "EndTime": "終了時刻",
        "FilePath": "ファイルパス",
        "TargetSheet": "転記シート",
        "SkipDownload": "DLスキップ",
        "SearchKey": "検索キー",
        "DownloadURL": "URL",
        "ActionAfter": "完了後動作",
        "CloseAfter": "終了後閉じる",
        "PopupMessage": "メッセージ",
        "MacroName": "マクロ名"
    }

    try:
        wb = openpyxl.load_workbook(target_file)
        if "TaskList" in wb.sheetnames:
            ws = wb["TaskList"]
            
            # ヘッダー行(1行目)を走査して置換
            for cell in ws[1]:
                if cell.value in mapping:
                    print(f"Renaming {cell.value} -> {mapping[cell.value]}")
                    cell.value = mapping[cell.value]
            
            wb.save(target_file)
            print("Migration completed successfully.")
        else:
            print("Sheet 'TaskList' not found.")
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    migrate_headers()
